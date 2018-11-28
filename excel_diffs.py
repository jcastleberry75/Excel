#!/usr/bin/env python3

__author__ = "Your Name"
__version__ = "0.1.0"
__license__ = "MIT"


import os
from colorama import *
from openpyxl import load_workbook


def main():
    import os
    from pathlib import Path
    import sys
    from time import sleep
    from colorama import init, Fore, Style
    import platform



    # Initialize and define colors
    def colors():
        init(autoreset=True)
        global yellow
        yellow = Style.BRIGHT + Fore.YELLOW
        global cyan
        cyan = Style.BRIGHT + Fore.CYAN
        global red
        red = Style.BRIGHT + Fore.RED
        global green
        green = Style.BRIGHT + Fore.GREEN
        global white
        white = Style.BRIGHT + Fore.WHITE

    def clear_screen():
        found_os = platform.system()
        win_os = 'Windows'
        if found_os is win_os:
            os.system('cls')
        else:
            os.system('clear')


    colors()
    clear_screen()

    print()
    print(cyan + "XXXXXXXXXXXXXXX  EXCEL DIFF GENERATOR XXXXXXXXXXXXXXX")
    print('\n')
    excel_datasheet = 'diff_data.xlsx'
    file_found = Path(excel_datasheet).exists()
    if file_found is True:
        pass
    else:
        print(red + "No Excel Data Sheet Found... Exiting in 5 Seconds")
        sleep(5)
        sys.exit()

    diff_data = load_workbook(excel_datasheet)
    print("Parsing the following sheets: ")
    print(diff_data.sheetnames[0:2])
    data1 = diff_data['data1']
    data2 = diff_data['data2']
    print()

    def dif_parser(sheet1,sheet2):
        sheet1_data_found = []
        sheet2_data_found = []
        data1_not_found_data2 = []
        data2_not_found_data1 = []
        sheet1_rows_total = sheet1.max_row
        sheet2_rows_total = sheet2.max_row
        print((str(sheet1_rows_total) + " rows found in " + str(sheet1)))
        print((str(sheet2_rows_total) + " rows found in " + str(sheet2)))
        print()
        print("Found the following cell data in sheet data1:")
        print()
        for rowOfCellObjects in sheet1['A1':sheet1.max_row]:
            for cellObj in rowOfCellObjects:
                print(cyan + cellObj.coordinate + ": ", cellObj.value)
                readable_data1 = str(cellObj.value)
                sheet1_data_found.append(readable_data1)
        clear_screen()
        print()
        for rowOfCellObjects in sheet2['A1':sheet1.max_row]:
            for cellObj in rowOfCellObjects:
                print(cyan + cellObj.coordinate + ": ", cellObj.value)
                readable_data2 = str(cellObj.value)
                sheet2_data_found.append(readable_data2)
        for item in sheet1_data_found:
            if item not in sheet2_data_found:
                data1_not_found_data2.append(item)
            else:
                pass
        for item in sheet2_data_found:
            if item not in sheet1_data_found:
                data2_not_found_data1.append(item)
            else:
                pass
        print(data1_not_found_data2)
        dat1_not_dat2 = open('dat1_not_dat2.txt', 'w')
        for item in data1_not_found_data2:
            dat1_not_dat2.write(item + '\n')
        print(data2_not_found_data1)
        dat2_not_dat1 = open('dat2_not_dat1.txt', 'w')
        for item in data2_not_found_data1:
            dat2_not_dat1.write(item + '\n')


    dif_parser(data1, data2)


if __name__ == "__main__":
    main()
